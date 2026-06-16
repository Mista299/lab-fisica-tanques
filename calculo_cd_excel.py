import numpy as np
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

g = 9.81

cuadrado = {
    "nombre": "Tanque cuadrado",
    "h0": 0.25, "AT": 0.0056, "A0": 0.000025,
    "fluidos": [
        {"nombre": "Agua",   "path": "datos_experimentales/cuadrado/Datos agua.txt"},
        {"nombre": "Aceite", "path": "datos_experimentales/cuadrado/Datos aceite.txt"},
    ],
}

R_b = 0.057; R_t = 0.0615; h0_j = 0.113
alpha_j = (R_t - R_b) / h0_j
A0_j = math.pi * 0.0025**2

jarra = {
    "nombre": "Jarra (tronco de cono)",
    "h0": h0_j, "AT": None, "A0": A0_j,
    "fluidos": [
        {"nombre": "Agua",   "path": "datos_experimentales/tronco_de_cono/Datos Agua.txt"},
        {"nombre": "Aceite", "path": "datos_experimentales/tronco_de_cono/Datos aceite.txt"},
    ],
}

def parse_datos(filepath):
    t_data, y_data = [], []
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line: continue
            parts = line.replace(',', '.').split('\t')
            if len(parts) >= 3:
                try:
                    t_data.append(float(parts[0]))
                    y_data.append(float(parts[2]))
                except ValueError: continue
    return np.array(t_data), np.array(y_data)

def calc_cd_cuadrado(t, h, h0, AT, A0):
    if h <= 0 or h >= h0 or t <= 0: return None
    k = (math.sqrt(h0) - math.sqrt(h)) / t
    return (k * 2 * AT) / (A0 * math.sqrt(2 * g))

def f_cono(h):
    return (R_b**2 * (math.sqrt(h0_j) - math.sqrt(h))
            + (2/3) * R_b * alpha_j * (h0_j**1.5 - h**1.5)
            + (1/5) * alpha_j**2 * (h0_j**2.5 - h**2.5))

def calc_cd_cono(t, h, h0, A0):
    if h <= 0 or h >= h0 or t <= 0: return None
    f = f_cono(h)
    return (2 * math.pi / (t * A0 * math.sqrt(2 * g))) * f

# Estilos
hdr_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=13)
fluid_font = Font(bold=True, size=12, color="003366")
thin_bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
c_align = Alignment(horizontal='center', vertical='center')

def style_hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font = hdr_font; cell.alignment = c_align; cell.border = thin_bd

def style_row(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = c_align; cell.border = thin_bd

wb = Workbook()
wb.remove(wb.active)

for config in [cuadrado, jarra]:
    ws = wb.create_sheet(title=config["nombre"][:31])
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=f"Calculo del coeficiente de descarga - {config['nombre']}").font = title_font
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    if config["AT"] is not None:
        ws.cell(row=row, column=1, value=f"h0 = {config['h0']} m    AT = {config['AT']} m2    A0 = {config['A0']} m2    g = {g} m/s2")
    else:
        ws.cell(row=row, column=1, value=f"h0 = {config['h0']} m    Rb = {R_b} m    Rt = {R_t} m    A0 = {A0_j:.6f} m2    g = {g} m/s2")
    row += 2

    for fluido in config["fluidos"]:
        t, y = parse_datos(fluido["path"])
        resultados = []
        for i in range(len(t)):
            if config["AT"] is not None:
                cd = calc_cd_cuadrado(t[i], y[i], config["h0"], config["AT"], config["A0"])
            else:
                cd = calc_cd_cono(t[i], y[i], config["h0"], config["A0"])
            if cd is not None and 0 < cd < 2:
                resultados.append((t[i], y[i], cd))

        cd_arr = np.array([r[2] for r in resultados])
        N = len(cd_arr)
        cd_mean = np.mean(cd_arr)
        cd_std = np.std(cd_arr, ddof=1)
        cd_uncert = cd_std / math.sqrt(N)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Fluido: {fluido['nombre']}").font = fluid_font
        row += 1

        headers = ["#", "t (s)", "h (m)", "Cd", "|Cd - Cd prom|"]
        for j, hdr in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=hdr)
        style_hdr(ws, row, len(headers))
        row += 1

        for i, (tv, hv, cdv) in enumerate(resultados):
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=round(tv, 3))
            ws.cell(row=row, column=3, value=round(hv, 5))
            ws.cell(row=row, column=4, value=round(cdv, 4))
            ws.cell(row=row, column=5, value=round(abs(cdv - cd_mean), 4))
            style_row(ws, row, len(headers))
            row += 1

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="Estadistica").font = Font(bold=True, size=11)
        row += 1

        stats = [
            ("N (puntos validos)", N),
            ("Cd promedio", cd_mean),
            ("Desviacion estandar (sigma)", cd_std),
            ("Incertidumbre (sigma/raiz(N))", cd_uncert),
            ("Resultado final", f"Cd = {cd_mean:.4f} ± {cd_uncert:.4f}"),
        ]
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value if isinstance(value, str) else round(value, 4))
            row += 1

        row += 2

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions['A'].width = 28

# ============================================================
# Hoja: Incertidumbres
# ============================================================
ws = wb.create_sheet("Incertidumbres")
row = 1
SEP = ","  # coma para Google Sheets

def write_param(ws, r, c, label, value, unit=""):
    ws.cell(row=r, column=c, value=label).font = Font(bold=True)
    ws.cell(row=r, column=c+1, value=value)
    if unit:
        ws.cell(row=r, column=c+2, value=unit)

def write_formula(ws, r, c, label, formula):
    ws.cell(row=r, column=c, value=label).font = Font(bold=True)
    ws.cell(row=r, column=c+1, value=formula)

# ============================================================
# SECCION 1: TANQUE CUADRADO
# ============================================================
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws.cell(row=row, column=1, value="Incertidumbre - Tanque cuadrado").font = Font(bold=True, size=13)
row += 2

# --- Parametros ---
ws.cell(row=row, column=1, value="Parametros geometricos:").font = Font(bold=True, size=11)
row += 1
params_q = [
    ("a (m)", 0.07, 0.001, "Regla"),
    ("b (m)", 0.08, 0.001, "Regla"),
    ("AT = a*b (m2)", None, None, ""),
    ("d_orificio (m)", round(math.sqrt(4*0.000025/math.pi),6), 0.00005, "Calibre"),
    ("A0 (m2)", 0.000025, None, ""),
    ("h0 (m)", 0.25, 0.001, "Regla"),
]
# Headers
for j, hdr in enumerate(["Parametro", "Valor", "Incertidumbre (±)", "Instrumento"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# Write params with formulas where applicable
ws.cell(row=row, column=1, value="a (m)"); ws.cell(row=row, column=2, value=0.07)
ws.cell(row=row, column=3, value=0.001); ws.cell(row=row, column=4, value="Regla")
r_a, r_b = row, row+1
row += 1
ws.cell(row=row, column=1, value="b (m)"); ws.cell(row=row, column=2, value=0.08)
ws.cell(row=row, column=3, value=0.001); ws.cell(row=row, column=4, value="Regla")
row += 1
# AT = a*b formula
ws.cell(row=row, column=1, value="AT = a*b (m2)")
ws.cell(row=row, column=2, value=f"=B{r_a}*B{r_b}")
row += 1
# d_orificio
ws.cell(row=row, column=1, value="d_orificio (m)")
d_orif_val = math.sqrt(4*0.000025/math.pi)
ws.cell(row=row, column=2, value=round(d_orif_val, 6))
ws.cell(row=row, column=3, value=0.00005); ws.cell(row=row, column=4, value="Calibre")
r_dorif = row
row += 1
# A0
ws.cell(row=row, column=1, value="A0 = pi*d^2/4 (m2)")
ws.cell(row=row, column=2, value=f"=PI()*B{r_dorif}^2/4")
r_A0q = row
row += 1
# h0
ws.cell(row=row, column=1, value="h0 (m)"); ws.cell(row=row, column=2, value=0.25)
ws.cell(row=row, column=3, value=0.001); ws.cell(row=row, column=4, value="Regla")
r_h0q = row
row += 2

# --- Incertidumbre relativa tipo B ---
ws.cell(row=row, column=1, value="Incertidumbre instrumental (tipo B):").font = Font(bold=True, size=11)
row += 1
for j, hdr in enumerate(["Contribucion", "Formula", "Valor (%)"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# u_AT_rel = SQRT((da/a)^2 + (db/b)^2)
ws.cell(row=row, column=1, value="u(AT)/AT")
ws.cell(row=row, column=2, value=f"=SQRT((C{r_a}/B{r_a})^2 + (C{r_b}/B{r_b})^2)")
ws.cell(row=row, column=3, value=f"=100*B{row}")
r_uAT = row; row += 1

# u_A0_rel = 2*dd/d
ws.cell(row=row, column=1, value="u(A0)/A0")
ws.cell(row=row, column=2, value=f"=2*C{r_dorif}/B{r_dorif}")
ws.cell(row=row, column=3, value=f"=100*B{row}")
r_uA0 = row; row += 1

# u_h0_contrib
ws.cell(row=row, column=1, value="u(h0)/Cd (contrib)")
ws.cell(row=row, column=2, value=f"=C{r_h0q}/(2*SQRT(B{r_h0q})*(SQRT(B{r_h0q})-SQRT(B{r_h0q}/2)))")
ws.cell(row=row, column=3, value=f"=100*B{row}")
r_uh0q = row; row += 1

# u_sys_rel = SQRT(uAT^2 + uA0^2 + uh0^2)
ws.cell(row=row, column=1, value="u_sys/Cd (combinada)")
ws.cell(row=row, column=2, value=f"=SQRT(B{r_uAT}^2 + B{r_uA0}^2 + B{r_uh0q}^2)")
ws.cell(row=row, column=3, value=f"=100*B{row}")
r_usysq = row; row += 2

# --- Resultados combinados ---
ws.cell(row=row, column=1, value="Resultados combinados:").font = Font(bold=True, size=11)
row += 1
for j, hdr in enumerate(["Fluido", "Cd", "u_tipo_A (estad)", "u_tipo_B (instr)", "u_combinada", "Resultado final"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# Agua
r_cd_agua_q = row
ws.cell(row=row, column=1, value="Agua"); ws.cell(row=row, column=2, value=0.4164)
ws.cell(row=row, column=3, value=0.0011)
ws.cell(row=row, column=4, value=f"=B{r_cd_agua_q}*B{r_usysq}")
ws.cell(row=row, column=5, value=f"=SQRT(C{r_cd_agua_q}^2 + D{r_cd_agua_q}^2)")
ws.cell(row=row, column=6, value=f"=B{r_cd_agua_q} & \" ± \" & ROUND(E{r_cd_agua_q}, 4)")
row += 1

# Aceite
r_cd_ac_q = row
ws.cell(row=row, column=1, value="Aceite"); ws.cell(row=row, column=2, value=0.3983)
ws.cell(row=row, column=3, value=0.0008)
ws.cell(row=row, column=4, value=f"=B{r_cd_ac_q}*B{r_usysq}")
ws.cell(row=row, column=5, value=f"=SQRT(C{r_cd_ac_q}^2 + D{r_cd_ac_q}^2)")
ws.cell(row=row, column=6, value=f"=B{r_cd_ac_q} & \" ± \" & ROUND(E{r_cd_ac_q}, 4)")
row += 3

# ============================================================
# SECCION 2: JARRA TRONCOCONICA
# ============================================================
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws.cell(row=row, column=1, value="Incertidumbre - Jarra troncoconica").font = Font(bold=True, size=13)
row += 2

ws.cell(row=row, column=1, value="Parametros geometricos:").font = Font(bold=True, size=11)
row += 1
for j, hdr in enumerate(["Parametro", "Valor", "Incertidumbre (±)", "Instrumento"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# Rb
ws.cell(row=row, column=1, value="Rb (m)"); ws.cell(row=row, column=2, value=0.057)
ws.cell(row=row, column=3, value=0.0005); ws.cell(row=row, column=4, value="Calibre")
r_Rb = row; row += 1
# Rt
ws.cell(row=row, column=1, value="Rt (m)"); ws.cell(row=row, column=2, value=0.0615)
ws.cell(row=row, column=3, value=0.0005); ws.cell(row=row, column=4, value="Calibre")
r_Rt = row; row += 1
# h0
ws.cell(row=row, column=1, value="h0 (m)"); ws.cell(row=row, column=2, value=0.113)
ws.cell(row=row, column=3, value=0.001); ws.cell(row=row, column=4, value="Regla")
r_h0j = row; row += 1
# d_orificio
ws.cell(row=row, column=1, value="d_orificio (m)"); ws.cell(row=row, column=2, value=0.005)
ws.cell(row=row, column=3, value=0.00005); ws.cell(row=row, column=4, value="Calibre")
r_dj = row; row += 1
# A0
ws.cell(row=row, column=1, value="A0 = pi*d^2/4 (m2)")
ws.cell(row=row, column=2, value=f"=PI()*B{r_dj}^2/4")
r_A0j = row; row += 2

# --- Incertidumbre tipo B (derivadas numericas precalculadas) ---
ws.cell(row=row, column=1, value="Incertidumbre instrumental (tipo B):").font = Font(bold=True, size=11)
row += 1
ws.cell(row=row, column=1, value="Nota: contribuciones obtenidas por derivacion numerica (diferencia central, paso 0.1%)").font = Font(italic=True, size=9)
row += 1
for j, hdr in enumerate(["Contribucion", "Valor relativo (%)"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# Rb contrib
ws.cell(row=row, column=1, value="u(Rb)/Cd"); ws.cell(row=row, column=2, value=0.0044)
r_uRb = row; row += 1
# Rt contrib
ws.cell(row=row, column=1, value="u(Rt)/Cd"); ws.cell(row=row, column=2, value=0.0122)
r_uRt = row; row += 1
# h0 contrib
ws.cell(row=row, column=1, value="u(h0)/Cd"); ws.cell(row=row, column=2, value=0.0147)
r_uh0j = row; row += 1
# A0 contrib
ws.cell(row=row, column=1, value="u(A0)/A0")
ws.cell(row=row, column=2, value=f"=2*C{r_dj}/B{r_dj}")
r_uA0j = row; row += 1

# u_sys_rel
ws.cell(row=row, column=1, value="u_sys/Cd (combinada)")
ws.cell(row=row, column=2, value=f"=SQRT(B{r_uRb}^2 + B{r_uRt}^2 + B{r_uh0j}^2 + B{r_uA0j}^2)")
r_usysj = row; row += 2

# --- Resultados ---
ws.cell(row=row, column=1, value="Resultados combinados:").font = Font(bold=True, size=11)
row += 1
for j, hdr in enumerate(["Fluido", "Cd", "u_tipo_A (estad)", "u_tipo_B (instr)", "u_combinada", "Resultado final"], 1):
    ws.cell(row=row, column=j, value=hdr).font = Font(bold=True)
row += 1

# Agua
r_cd_a_j = row
ws.cell(row=row, column=1, value="Agua"); ws.cell(row=row, column=2, value=0.5899)
ws.cell(row=row, column=3, value=0.0008)
ws.cell(row=row, column=4, value=f"=B{r_cd_a_j}*B{r_usysj}")
ws.cell(row=row, column=5, value=f"=SQRT(C{r_cd_a_j}^2 + D{r_cd_a_j}^2)")
ws.cell(row=row, column=6, value=f"=B{r_cd_a_j} & \" ± \" & ROUND(E{r_cd_a_j}, 4)")
row += 1

# Aceite
r_cd_o_j = row
ws.cell(row=row, column=1, value="Aceite"); ws.cell(row=row, column=2, value=0.4446)
ws.cell(row=row, column=3, value=0.0009)
ws.cell(row=row, column=4, value=f"=B{r_cd_o_j}*B{r_usysj}")
ws.cell(row=row, column=5, value=f"=SQRT(C{r_cd_o_j}^2 + D{r_cd_o_j}^2)")
ws.cell(row=row, column=6, value=f"=B{r_cd_o_j} & \" ± \" & ROUND(E{r_cd_o_j}, 4)")

# Ajustar ancho
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 22

wb.save("calculo_cd_experimental.xlsx")
print("Excel con valores calculados generado: calculo_cd_experimental.xlsx")
